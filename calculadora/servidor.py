#!/usr/bin/env python3
"""
Calculadora de Importacion v2 - Paracarpinteros
Servidor local con:
  - Extraccion de imagenes embebidas de Excel (openpyxl)
  - Base de datos SQLite para pedidos por proveedor
  - Conexion API Odoo 19 para ventas/stock
  - Generador de Excel de reposicion con imagenes

Uso: python servidor.py
"""

import os
import sys
import json
import base64
import io
import webbrowser
import threading
import sqlite3
import tempfile
import zipfile
import re
import time
from datetime import datetime
from http.server import HTTPServer, SimpleHTTPRequestHandler
from urllib.parse import parse_qs, urlparse
from pathlib import Path

# ==================== DEPENDENCIAS ====================
def install_if_missing(pkg, import_name=None):
    try:
        __import__(import_name or pkg)
    except ImportError:
        print(f"  Instalando {pkg}...")
        os.system(f'"{sys.executable}" -m pip install {pkg} --quiet')

install_if_missing('openpyxl')
install_if_missing('Pillow', 'PIL')
install_if_missing('xlsxwriter')
install_if_missing('requests')
install_if_missing('beautifulsoup4', 'bs4')

import openpyxl
from PIL import Image as PILImage
import xlsxwriter
import requests
from bs4 import BeautifulSoup

# ==================== CONFIG ====================
PORT = int(os.environ.get('PORT', 5555))
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.environ.get('DB_PATH', os.path.join(SCRIPT_DIR, 'pedidos.db'))

# ==================== SCRAPER PARACARPINTEROS ====================
_scrape_cache = {}  # url -> data

def scrape_paracarpinteros(url):
    """Extrae imagen, nombre, referencia, precio y SKU de una URL de paracarpinteros.com"""
    if url in _scrape_cache:
        return _scrape_cache[url]

    result = {'nombre': '', 'referencia': '', 'sku': '', 'imagen_url': '',
              'imagen_b64': '', 'precio_crc': 0, 'categoria': ''}

    try:
        url = url.strip()
        if not url.startswith('http'):
            url = 'https://' + url
        if 'paracarpinteros.com' not in url:
            return result

        print(f"  [SCRAPE] Visitando {url[:80]}...")
        headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/120'}
        resp = requests.get(url, headers=headers, timeout=15)
        if resp.status_code != 200:
            print(f"  [SCRAPE] HTTP {resp.status_code}")
            return result

        soup = BeautifulSoup(resp.text, 'html.parser')

        # Nombre del producto (h1 o og:title)
        h1 = soup.find('h1')
        if h1:
            result['nombre'] = h1.get_text(strip=True)
        else:
            og = soup.find('meta', property='og:title')
            if og:
                result['nombre'] = og.get('content', '')

        # Referencia: buscar Ref.XXX en el nombre
        nombre = result['nombre']
        ref_match = re.search(r'Ref\.?\s*([A-Z]\d{1,3})', nombre, re.IGNORECASE)
        if ref_match:
            result['referencia'] = ref_match.group(1).upper()

        # SKU: buscar [AXXX] en la pagina
        sku_match = re.search(r'\[([A-Z]\d{2,5})\]', resp.text)
        if sku_match:
            result['sku'] = sku_match.group(1)

        # Precio
        price_el = soup.find('span', class_=re.compile(r'oe_price|oe_currency_value|product_price'))
        if price_el:
            price_text = price_el.get_text(strip=True).replace(',', '').replace('₡', '').strip()
            try:
                result['precio_crc'] = float(re.sub(r'[^\d.]', '', price_text))
            except:
                pass
        # Fallback: buscar precio en meta o span
        if not result['precio_crc']:
            for span in soup.find_all('span'):
                txt = span.get_text(strip=True)
                if '₡' in txt or 'CRC' in txt:
                    m = re.search(r'[\d,.]+', txt.replace(',', ''))
                    if m:
                        try:
                            result['precio_crc'] = float(m.group().replace(',', ''))
                            break
                        except:
                            pass

        # Categoria
        breadcrumb = soup.find('ol', class_='breadcrumb')
        if breadcrumb:
            items = breadcrumb.find_all('li')
            if len(items) >= 2:
                result['categoria'] = items[-2].get_text(strip=True)

        # Imagen principal del producto
        img_url = ''
        # Buscar imagen og:image (suele ser la principal)
        og_img = soup.find('meta', property='og:image')
        if og_img:
            img_url = og_img.get('content', '')

        # Si no, buscar en el carousel de producto
        if not img_url:
            product_img = soup.find('img', class_=re.compile(r'product|js_variant_img'))
            if product_img:
                img_url = product_img.get('src', '') or product_img.get('data-src', '')

        # Si no, buscar /web/image/product
        if not img_url:
            for img in soup.find_all('img'):
                src = img.get('src', '') or img.get('data-src', '')
                if '/web/image/product' in src:
                    img_url = src
                    break

        if img_url:
            if img_url.startswith('/'):
                img_url = 'https://www.paracarpinteros.com' + img_url
            result['imagen_url'] = img_url

            # Descargar y convertir a base64
            try:
                img_resp = requests.get(img_url, headers=headers, timeout=10)
                if img_resp.status_code == 200:
                    result['imagen_b64'] = img_to_b64(img_resp.content, max_size=150)
                    print(f"  [SCRAPE] OK: {result['referencia'] or result['sku']} - {result['nombre'][:40]}")
            except Exception as e:
                print(f"  [SCRAPE] Error descargando imagen: {e}")

    except Exception as e:
        print(f"  [SCRAPE] Error: {e}")

    _scrape_cache[url] = result
    return result

def scrape_multiple_urls(url_list):
    """Scrape varias URLs en secuencia, retorna dict {url: data}"""
    results = {}
    total = len(url_list)
    for i, url in enumerate(url_list):
        print(f"  [SCRAPE] {i+1}/{total}")
        results[url] = scrape_paracarpinteros(url)
    return results

def scrape_supplier_title(url):
    """Extrae solo el título de una URL de proveedor (Alibaba, AliExpress, 1688)"""
    try:
        url = url.strip()
        if not url.startswith('http'):
            url = 'https://' + url
        headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/120',
                   'Accept-Language': 'es-ES,es;q=0.9'}
        print(f"  [TITLE] Fetching {url[:80]}...")
        resp = requests.get(url, headers=headers, timeout=15)
        if resp.status_code != 200:
            return ''
        soup = BeautifulSoup(resp.text, 'html.parser')
        # Intentar h1 primero, luego <title>, luego og:title
        h1 = soup.find('h1')
        if h1 and len(h1.get_text(strip=True)) > 5:
            return h1.get_text(strip=True)
        title_tag = soup.find('title')
        if title_tag:
            t = title_tag.get_text(strip=True)
            # Limpiar sufijos de plataforma
            for suf in [' - Alibaba.com', ' | Alibaba', ' - AliExpress', ' - 1688.com', '- Compra']:
                t = t.split(suf)[0]
            return t.strip()
        og = soup.find('meta', property='og:title')
        if og:
            return og.get('content', '').strip()
    except Exception as e:
        print(f"  [TITLE] Error: {e}")
    return ''

# ==================== BASE DE DATOS ====================
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")  # Mejor para Google Drive
    conn.execute("PRAGMA busy_timeout=5000")
    return conn

def init_db():
    conn = get_db()
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS proveedores (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT UNIQUE NOT NULL,
            contacto TEXT DEFAULT '',
            notas TEXT DEFAULT '',
            created_at TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS pedidos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            proveedor_id INTEGER NOT NULL,
            titulo TEXT NOT NULL,
            estado TEXT DEFAULT 'borrador',
            fecha TEXT DEFAULT (datetime('now')),
            notas TEXT DEFAULT '',
            flete REAL DEFAULT 0,
            seguro REAL DEFAULT 0,
            gastos_json TEXT DEFAULT '[]',
            impuestos_json TEXT DEFAULT '[]',
            alloc_peso INTEGER DEFAULT 50,
            fx_usd_crc REAL DEFAULT 525,
            FOREIGN KEY (proveedor_id) REFERENCES proveedores(id)
        );
        CREATE TABLE IF NOT EXISTS productos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            pedido_id INTEGER NOT NULL,
            referencia TEXT DEFAULT '',
            nombre TEXT NOT NULL,
            variante TEXT DEFAULT '',
            url_proveedor TEXT DEFAULT '',
            url_paracarpinteros TEXT DEFAULT '',
            imagen_b64 TEXT DEFAULT '',
            moneda TEXT DEFAULT 'USD',
            precio REAL DEFAULT 0,
            cantidad REAL DEFAULT 0,
            peso_g REAL DEFAULT 0,
            vol_cm3 REAL DEFAULT 0,
            margen REAL DEFAULT 40,
            landed_unit REAL DEFAULT 0,
            pvp_unit REAL DEFAULT 0,
            odoo_product_id INTEGER DEFAULT NULL,
            FOREIGN KEY (pedido_id) REFERENCES pedidos(id) ON DELETE CASCADE
        );
        CREATE TABLE IF NOT EXISTS odoo_config (
            id INTEGER PRIMARY KEY CHECK (id = 1),
            url TEXT DEFAULT '',
            db_name TEXT DEFAULT '',
            username TEXT DEFAULT '',
            api_key TEXT DEFAULT '',
            last_sync TEXT DEFAULT ''
        );
        INSERT OR IGNORE INTO odoo_config (id) VALUES (1);

        CREATE TABLE IF NOT EXISTS lista_espera (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            referencia TEXT DEFAULT '',
            nombre_producto TEXT DEFAULT '',
            cliente TEXT NOT NULL,
            telefono TEXT DEFAULT '',
            pedido_id INTEGER DEFAULT NULL,
            estado TEXT DEFAULT 'esperando',
            fecha TEXT DEFAULT (datetime('now')),
            fecha_avisado TEXT DEFAULT '',
            notas TEXT DEFAULT ''
        );

        CREATE TABLE IF NOT EXISTS catalogo (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            referencia TEXT UNIQUE DEFAULT '',
            nombre TEXT NOT NULL,
            variante TEXT DEFAULT '',
            url_paracarpinteros TEXT DEFAULT '',
            url_proveedor TEXT DEFAULT '',
            imagen_b64 TEXT DEFAULT '',
            ultimo_precio REAL DEFAULT 0,
            ultimo_proveedor TEXT DEFAULT '',
            pvp_crc REAL DEFAULT 0,
            peso_g REAL DEFAULT 0,
            margen REAL DEFAULT 40,
            odoo_product_id INTEGER DEFAULT NULL,
            odoo_publicado INTEGER DEFAULT 0,
            veces_pedido INTEGER DEFAULT 0,
            created_at TEXT DEFAULT (datetime('now')),
            updated_at TEXT DEFAULT (datetime('now'))
        );
    """)
    # Migraciones de columnas en productos
    for col, default in [('nota', "''"), ('cliente_avisado', '0'), ('cliente_completado', '0')]:
        try:
            conn.execute(f"SELECT {col} FROM productos LIMIT 1")
        except:
            conn.execute(f"ALTER TABLE productos ADD COLUMN {col} DEFAULT {default}")
    conn.commit()
    conn.close()
    print("  [OK] Base de datos inicializada:", DB_PATH)

# ==================== EXTRACCION IMAGENES EXCEL ====================
def extract_images_from_xlsx(filepath):
    images = {}
    try:
        # Metodo principal: ZIP parsing (mas fiable)
        images = extract_images_from_zip(filepath)
        if not images:
            # Fallback: openpyxl _images
            wb = openpyxl.load_workbook(filepath)
            ws = wb.active
            if hasattr(ws, '_images') and ws._images:
                for img in ws._images:
                    try:
                        anchor = img.anchor
                        row = None
                        if hasattr(anchor, '_from') and anchor._from:
                            row = anchor._from.row
                        if row is not None:
                            img_data = img._data() if hasattr(img, '_data') else None
                            if img_data:
                                images[row] = img_to_b64(img_data)
                    except:
                        pass
            wb.close()
    except Exception as e:
        print(f"  Error imagenes: {e}")
    return images

def extract_images_from_zip(filepath):
    images = {}
    try:
        with zipfile.ZipFile(filepath, 'r') as z:
            media_files = sorted([f for f in z.namelist() if f.startswith('xl/media/')])
            if not media_files:
                return images

            drawing_rels = {}
            for name in z.namelist():
                if 'drawing' in name and name.endswith('.rels'):
                    content = z.read(name).decode('utf-8', errors='ignore')
                    for m in re.finditer(r'Id="(rId\d+)"[^>]*Target="([^"]*)"', content):
                        rid, target = m.groups()
                        drawing_rels[rid] = target.replace('../', 'xl/')

            anchor_map = {}
            for name in z.namelist():
                if 'drawing' in name and name.endswith('.xml') and '.rels' not in name:
                    content = z.read(name).decode('utf-8', errors='ignore')
                    blocks = re.split(r'<xdr:twoCellAnchor|<xdr:oneCellAnchor', content)
                    for block in blocks[1:]:
                        from_match = re.search(r'<xdr:from>.*?<xdr:row>(\d+)</xdr:row>', block, re.DOTALL)
                        to_match = re.search(r'<xdr:to>.*?<xdr:row>(\d+)</xdr:row>', block, re.DOTALL)
                        rid_match = re.search(r'r:embed="(rId\d+)"', block)
                        if from_match and rid_match:
                            from_row = int(from_match.group(1))
                            # Usar el centro entre from y to para compensar imágenes desplazadas
                            if to_match:
                                to_row = int(to_match.group(1))
                                row = (from_row + to_row) // 2 if to_row > from_row else from_row
                            else:
                                row = from_row
                            rid = rid_match.group(1)
                            if rid in drawing_rels:
                                anchor_map[row] = drawing_rels[rid]

            if anchor_map:
                for row, media_path in anchor_map.items():
                    for mf in media_files:
                        if mf.endswith(os.path.basename(media_path)):
                            images[row] = img_to_b64(z.read(mf))
                            break
            else:
                for idx, mf in enumerate(media_files):
                    images[idx + 1] = img_to_b64(z.read(mf))
    except Exception as e:
        print(f"  Error ZIP: {e}")
    return images

def img_to_b64(data, max_size=120):
    try:
        pil = PILImage.open(io.BytesIO(data))
        pil.thumbnail((max_size, max_size), PILImage.LANCZOS)
        buf = io.BytesIO()
        pil.save(buf, format='PNG')
        return "data:image/png;base64," + base64.b64encode(buf.getvalue()).decode()
    except:
        return "data:image/png;base64," + base64.b64encode(data).decode()

def parse_excel(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    all_rows = []
    for row in ws.iter_rows(values_only=True):
        all_rows.append([str(c) if c is not None else '' for c in row])
    wb.close()

    if not all_rows:
        return [], [], {}

    # Detectar fila de headers: la fila donde MAS celdas individuales matchean keywords
    header_keywords = ['nombre','precio','ref','producto','cantidad','qty','price','descripcion',
                       'variante','peso','weight','url','unit','total','color','size','sku']
    header_idx = 0
    best_score = 0
    for i, row in enumerate(all_rows[:15]):
        # Contar cuantas celdas distintas contienen un keyword (celdas cortas = probables headers)
        score = 0
        for cell in row:
            c = cell.lower().strip()
            if not c or len(c) > 40:  # headers suelen ser cortos
                continue
            if any(kw in c for kw in header_keywords):
                score += 1
        if score > best_score:
            best_score = score
            header_idx = i

    headers = all_rows[header_idx]
    data_rows = all_rows[header_idx + 1:]
    images = extract_images_from_xlsx(filepath)

    adjusted = {}
    for row_idx, img_data in images.items():
        data_idx = row_idx - header_idx - 1
        if 0 <= data_idx < len(data_rows):
            adjusted[data_idx] = img_data

    return headers, data_rows, adjusted

# ==================== ODOO API ====================
def odoo_connect():
    conn = get_db()
    cfg = conn.execute("SELECT * FROM odoo_config WHERE id=1").fetchone()
    conn.close()

    if not cfg:
        print("  [ODOO] No hay fila en odoo_config")
        return None, "Odoo no configurado - no hay registro en BD"

    cfg_dict = dict(cfg)
    print(f"  [ODOO] Config leida: url='{cfg_dict.get('url','')}' db='{cfg_dict.get('db_name','')}' user='{cfg_dict.get('username','')}' key={'***' if cfg_dict.get('api_key') else 'VACIA'}")

    if not cfg_dict.get('url') or not cfg_dict.get('api_key'):
        return None, f"Odoo no configurado - url={'OK' if cfg_dict.get('url') else 'VACIA'}, key={'OK' if cfg_dict.get('api_key') else 'VACIA'}"

    try:
        import xmlrpc.client
        url = cfg_dict['url'].rstrip('/')
        print(f"  [ODOO] Conectando a {url}/xmlrpc/2/common ...")
        common = xmlrpc.client.ServerProxy(f'{url}/xmlrpc/2/common')
        uid = common.authenticate(cfg_dict['db_name'], cfg_dict['username'], cfg_dict['api_key'], {})
        if not uid:
            return None, "Credenciales invalidas - authenticate devolvio False"
        print(f"  [ODOO] Conectado! uid={uid}")
        models = xmlrpc.client.ServerProxy(f'{url}/xmlrpc/2/object')
        return {'uid': uid, 'models': models, 'db': cfg_dict['db_name'],
                'password': cfg_dict['api_key'], 'url': url}, None
    except Exception as e:
        print(f"  [ODOO] Error conexion: {e}")
        return None, str(e)

def clean_ref(ref):
    """Limpia referencia: 'Ref.P07' -> 'P07', 'Ref: A476' -> 'A476'"""
    r = re.sub(r'^Ref[\.:;]?\s*', '', ref.strip(), flags=re.IGNORECASE)
    return r.strip()

def odoo_get_stock_and_orders(odoo, ref_list):
    """Consulta stock, ventas y pedidos de compra abiertos por referencia"""
    try:
        m = odoo['models']
        db, uid, pwd = odoo['db'], odoo['uid'], odoo['password']

        # Limpiar referencias: "Ref.P07" -> "P07"
        clean_refs = []
        ref_map = {}  # clean -> original
        for r in ref_list:
            c = clean_ref(r)
            if c:
                clean_refs.append(c)
                ref_map[c] = r

        if not clean_refs:
            return {}, None

        print(f"  [ODOO] Buscando refs: {clean_refs[:10]}...")

        # Buscar por default_code exacto
        product_ids = m.execute_kw(db, uid, pwd, 'product.product', 'search_read',
            [[['default_code', 'in', clean_refs]]],
            {'fields': ['id', 'default_code', 'name', 'qty_available', 'virtual_available',
                         'incoming_qty']})

        # Refs no encontradas por default_code -> buscar por nombre o default_code parcial
        found_codes = {p['default_code'] for p in product_ids if p.get('default_code')}
        found_ids = {p['id'] for p in product_ids}
        missing = [r for r in clean_refs if r not in found_codes]

        if missing:
            print(f"  [ODOO] {len(missing)} no encontradas por default_code exacto, buscando por nombre/code parcial...")
            fields = ['id', 'default_code', 'name', 'qty_available', 'virtual_available', 'incoming_qty']
            # Buscar de a bloques de 10 para no saturar Odoo
            for i in range(0, len(missing), 10):
                batch = missing[i:i+10]
                # Dominio: (name ilike 'Ref.P07') OR (name ilike 'Ref.G08') OR (default_code ilike 'P07')...
                domain_parts = []
                for r in batch:
                    domain_parts.append(['name', 'ilike', f'Ref.{r}'])
                    domain_parts.append(['name', 'ilike', f'Ref {r}'])
                    domain_parts.append(['default_code', 'ilike', r])
                # Construir domain OR: n-1 operadores '|' para n condiciones
                domain = ['|'] * (len(domain_parts) - 1) + domain_parts
                try:
                    batch_results = m.execute_kw(db, uid, pwd, 'product.product', 'search_read',
                        [domain], {'fields': fields, 'limit': 100})
                    for p in batch_results:
                        if p['id'] not in found_ids:
                            product_ids.append(p)
                            found_ids.add(p['id'])
                except Exception as e:
                    print(f"  [ODOO] Error busqueda batch: {e}")

        results = {}
        pid_map = {}  # product_id -> original_ref
        seen_ids = set()
        for p in product_ids:
            if p['id'] in seen_ids:
                continue
            seen_ids.add(p['id'])
            dc = p.get('default_code') or ''
            name = p.get('name', '')

            # Determinar la ref original del Excel que corresponde a este producto
            original_ref = None
            # 1. Match por default_code limpio
            if dc in ref_map:
                original_ref = ref_map[dc]
            # 2. Match buscando en nombre: "Ref.P07" -> "P07" -> ref_map
            if not original_ref:
                for cr, orig in ref_map.items():
                    if cr.upper() in name.upper() or f'Ref.{cr}' in name or f'Ref {cr}' in name:
                        original_ref = orig
                        break
            # 3. Fallback: usar default_code tal cual
            if not original_ref:
                original_ref = dc or name[:30]

            pid_map[p['id']] = original_ref
            stock_data = {
                'odoo_id': p['id'],
                'name': name,
                'default_code': dc,
                'stock': p['qty_available'],          # stock fisico a mano
                'virtual': p['virtual_available'],     # stock pronosticado
                'incoming': p.get('incoming_qty', 0),  # en camino
                'sold_12m': 0,
                'avg_monthly': 0,
                'open_orders': []
            }
            results[original_ref] = stock_data
            print(f"  [ODOO] Match: '{original_ref}' -> '{name}' (dc={dc}, stock={p['qty_available']})")

        if not pid_map:
            return results, None

        all_pids = list(pid_map.keys())

        # Ventas ultimos 12 meses
        try:
            cutoff = (datetime.now().replace(year=datetime.now().year-1)).strftime('%Y-%m-%d')
            sale_lines = m.execute_kw(db, uid, pwd, 'sale.order.line', 'search_read',
                [[['product_id', 'in', all_pids],
                  ['state', 'in', ['sale', 'done']],
                  ['order_id.date_order', '>=', cutoff]]],
                {'fields': ['product_id', 'product_uom_qty']})
            for sl in sale_lines:
                pid = sl['product_id'][0] if isinstance(sl['product_id'], list) else sl['product_id']
                ref = pid_map.get(pid)
                if ref:
                    results[ref]['sold_12m'] += sl['product_uom_qty']
        except Exception as e:
            print(f"  [ODOO] Error ventas: {e}")

        # Pedidos de compra abiertos (purchase.order.line)
        try:
            po_lines = m.execute_kw(db, uid, pwd, 'purchase.order.line', 'search_read',
                [[['product_id', 'in', all_pids],
                  ['state', 'in', ['draft', 'sent', 'purchase']]]],
                {'fields': ['product_id', 'product_qty', 'price_unit', 'order_id',
                             'date_planned']})
            for pol in po_lines:
                pid = pol['product_id'][0] if isinstance(pol['product_id'], list) else pol['product_id']
                ref = pid_map.get(pid)
                if ref:
                    order_name = pol['order_id'][1] if isinstance(pol['order_id'], list) else str(pol['order_id'])
                    results[ref]['open_orders'].append({
                        'po': order_name,
                        'qty': pol['product_qty'],
                        'price': pol['price_unit'],
                        'date': pol.get('date_planned', '')
                    })
        except Exception as e:
            print(f"  [ODOO] Error pedidos compra: {e}")

        # Calcular media mensual
        for ref in results:
            s = results[ref]['sold_12m']
            results[ref]['avg_monthly'] = round(s / 12, 1) if s else 0

        return results, None
    except Exception as e:
        print(f"  [ODOO] Error general: {e}")
        return {}, str(e)

# ==================== PLANTILLA COTIZACION PROVEEDOR ====================
def generate_supplier_template(productos_json):
    """Genera Excel estandar para enviar al proveedor a cotizar"""
    productos = json.loads(productos_json) if isinstance(productos_json, str) else productos_json
    filepath = os.path.join(SCRIPT_DIR, 'cotizacion_proveedor.xlsx')

    wb = xlsxwriter.Workbook(filepath)
    ws = wb.add_worksheet('Cotizacion')

    # Formatos
    title_fmt = wb.add_format({'bold': True, 'font_size': 14, 'font_color': '#D97706',
                                'bottom': 2, 'bottom_color': '#D97706'})
    header_fmt = wb.add_format({'bold': True, 'bg_color': '#D97706', 'font_color': 'white',
                                 'border': 1, 'text_wrap': True, 'valign': 'vcenter', 'align': 'center'})
    cell_fmt = wb.add_format({'border': 1, 'valign': 'vcenter', 'text_wrap': True})
    cell_center = wb.add_format({'border': 1, 'valign': 'vcenter', 'align': 'center'})
    price_fmt = wb.add_format({'border': 1, 'valign': 'vcenter', 'num_format': '$#,##0.00',
                                'bg_color': '#FFF7ED'})
    qty_fmt = wb.add_format({'border': 1, 'valign': 'vcenter', 'num_format': '#,##0',
                              'bg_color': '#FFF7ED'})
    num_fmt = wb.add_format({'border': 1, 'valign': 'vcenter', 'num_format': '#,##0',
                              'align': 'center'})
    weight_total_fmt = wb.add_format({'border': 1, 'valign': 'vcenter', 'num_format': '#,##0',
                                       'align': 'center', 'bg_color': '#F0F9FF'})
    url_fmt = wb.add_format({'border': 1, 'valign': 'vcenter', 'font_color': '#2563EB',
                              'underline': True, 'font_size': 10})
    note_fmt = wb.add_format({'italic': True, 'font_color': '#888888', 'font_size': 10})
    total_fmt = wb.add_format({'bold': True, 'border': 1, 'num_format': '$#,##0.00',
                                'bg_color': '#FEF3C7'})
    total_num_fmt = wb.add_format({'bold': True, 'border': 1, 'num_format': '#,##0',
                                    'bg_color': '#FEF3C7', 'align': 'center'})
    subtotal_fmt = wb.add_format({'border': 1, 'valign': 'vcenter', 'num_format': '$#,##0.00',
                                   'bg_color': '#FFFBEB'})
    bold_r = wb.add_format({'bold': True, 'align': 'right', 'border': 1, 'valign': 'vcenter'})
    section_fmt = wb.add_format({'bold': True, 'font_size': 12, 'font_color': '#D97706',
                                  'bottom': 1, 'bottom_color': '#D97706'})
    nuevo_fmt = wb.add_format({'border': 1, 'valign': 'vcenter', 'align': 'center',
                                'bg_color': '#EDE9FE', 'font_color': '#7C3AED', 'bold': True})

    # Cell dimensions (pixels approx) for image scaling
    ROW_HEIGHT = 55  # pixels
    COL_IMG_WIDTH_PX = 85  # ~12 excel units ≈ 85px

    # Titulo
    ws.merge_range('A1:M1', 'SOLICITUD DE COTIZACION - Paracarpinteros', title_fmt)
    ws.write('A2', 'Fecha:', note_fmt)
    ws.write('B2', datetime.now().strftime('%d/%m/%Y'), note_fmt)
    ws.write('A3', 'Por favor complete las columnas marcadas en naranja claro', note_fmt)

    # Headers en fila 5 — Col: A=Img, B=Ref, C=Producto, D=Variante, E=URL, F=Cant, G=Precio, H=Subtotal, I=Peso(g), J=Peso Total(g), K=Disponible, L=Tiempo, M=Nuevo Odoo
    headers = ['', 'Ref', 'Producto', 'Descripcion/Variante', 'URL Referencia',
               'Cantidad', 'Precio Unit.', 'Subtotal', 'Peso unit.(g)', 'Peso Total(g)',
               'Disponible (S/N)', 'Tiempo entrega (dias)', 'Nuevo Odoo']
    ws.set_row(4, 30)
    for i, h in enumerate(headers):
        ws.write(4, i, h, header_fmt)

    ws.set_column(0, 0, 12)   # imagen
    ws.set_column(1, 1, 10)   # ref
    ws.set_column(2, 2, 35)   # nombre
    ws.set_column(3, 3, 20)   # variante
    ws.set_column(4, 4, 30)   # url
    ws.set_column(5, 5, 10)   # cantidad
    ws.set_column(6, 6, 12)   # precio
    ws.set_column(7, 7, 12)   # subtotal
    ws.set_column(8, 8, 11)   # peso unit
    ws.set_column(9, 9, 12)   # peso total
    ws.set_column(10, 10, 14)  # disponible
    ws.set_column(11, 11, 18)  # tiempo entrega
    ws.set_column(12, 12, 12)  # nuevo odoo

    row = 5
    tmp_files = []
    for p in productos:
        ws.set_row(row, ROW_HEIGHT)

        # Imagen — ajustar escala para que quepa en la celda
        if p.get('imgSrc', '').startswith('data:'):
            try:
                b64 = p['imgSrc'].split(',', 1)[1]
                img_bytes = base64.b64decode(b64)
                tmp = tempfile.NamedTemporaryFile(suffix='.png', delete=False)
                tmp.write(img_bytes)
                tmp.close()
                tmp_files.append(tmp.name)

                # Calcular escala real basada en dimensiones de imagen
                pil_img = PILImage.open(io.BytesIO(img_bytes))
                img_w, img_h = pil_img.size
                pil_img.close()

                # Escala para ajustar a celda (con margen de 4px)
                target_w = COL_IMG_WIDTH_PX - 4
                target_h = ROW_HEIGHT - 4
                scale_x = target_w / img_w if img_w > target_w else 1.0
                scale_y = target_h / img_h if img_h > target_h else 1.0
                scale = min(scale_x, scale_y)

                ws.insert_image(row, 0, tmp.name, {
                    'x_scale': scale, 'y_scale': scale,
                    'x_offset': 2, 'y_offset': 2,
                    'object_position': 1
                })
            except:
                pass

        ws.write(row, 1, p.get('referencia', ''), cell_center)
        ws.write(row, 2, p.get('nombre', ''), cell_fmt)
        ws.write(row, 3, p.get('variante', ''), cell_fmt)

        # URL de referencia — proveedor primero, luego paracarpinteros
        url = p.get('url_proveedor', '') or p.get('url_paracarpinteros', '')
        if url:
            ws.write_url(row, 4, url, url_fmt, 'Ver producto')
        else:
            ws.write(row, 4, '', cell_fmt)

        cant = p.get('cant', 0)
        peso_g = p.get('pesoG', 0) or 0

        ws.write(row, 5, cant, qty_fmt)
        ws.write(row, 6, '', price_fmt)          # VACIO - proveedor rellena
        # Subtotal = Cantidad * Precio (formula)
        xl_row = row + 1  # 1-based for formulas
        ws.write_formula(row, 7, f'=F{xl_row}*G{xl_row}', subtotal_fmt)
        ws.write(row, 8, peso_g, num_fmt)
        # Peso total = Cantidad * Peso unitario
        ws.write_formula(row, 9, f'=F{xl_row}*I{xl_row}', weight_total_fmt)
        ws.write(row, 10, '', cell_center)        # VACIO - proveedor rellena
        ws.write(row, 11, '', cell_center)         # VACIO - proveedor rellena
        # Nuevo Odoo — marcar si es producto nuevo
        is_nuevo = p.get('nuevo', False)
        ws.write(row, 12, 'SI' if is_nuevo else '', nuevo_fmt if is_nuevo else cell_center)
        row += 1

    # Fila TOTALES
    first_data = 6  # row 6 in Excel (1-based)
    last_data = row  # row after last product (0-based), in Excel = row (already 1-based for formula)

    ws.set_row(row, 25)
    ws.write(row, 4, 'TOTALES:', bold_r)
    ws.write_formula(row, 5, f'=SUM(F{first_data}:F{last_data})', total_num_fmt)
    ws.write_formula(row, 6, f'=SUM(G{first_data}:G{last_data})', total_fmt)
    ws.write_formula(row, 7, f'=SUM(H{first_data}:H{last_data})', total_fmt)
    ws.write_formula(row, 8, f'=SUM(I{first_data}:I{last_data})', total_num_fmt)
    ws.write_formula(row, 9, f'=SUM(J{first_data}:J{last_data})', total_num_fmt)
    row += 1

    # Fila COSTO ENVIO (proveedor rellena)
    ws.write(row, 4, 'COSTO ENVIO:', bold_r)
    ws.write(row, 7, '', price_fmt)  # proveedor rellena subtotal envio
    ship_row_xl = row + 1  # 1-based for formula
    row += 1

    # Fila TOTAL PEDIDO (productos + envio)
    totals_row_xl = row - 1  # the TOTALES row (1-based = row-1+1 = row... but totals is 2 rows up)
    ws.write(row, 4, 'TOTAL PEDIDO:', bold_r)
    ws.write_formula(row, 7, f'=H{totals_row_xl}+H{ship_row_xl}', total_fmt)
    row += 2

    # ============ OPCIONES DE ENVIO ============
    ws.merge_range(row, 0, row, 12, 'OPCIONES DE ENVIO (rellenar por proveedor)', section_fmt)
    row += 1

    # Headers envio
    ship_header = wb.add_format({'bold': True, 'bg_color': '#FEF3C7', 'border': 1,
                                  'valign': 'vcenter', 'align': 'center'})
    ship_labels = ['Metodo', 'Precio USD', 'Peso total (kg)', 'Tiempo estimado (dias)', 'Notas']
    for i, h in enumerate(ship_labels):
        ws.write(row, i, h, ship_header)
    row += 1

    # Filas de envio predefinidas
    ship_fill = wb.add_format({'border': 1, 'valign': 'vcenter', 'bg_color': '#FFF7ED'})
    for carrier in ['DHL', 'UPS', 'FEDEX', 'Maritimo', 'Otro']:
        ws.write(row, 0, carrier, cell_center)
        for c in range(1, 5):
            ws.write(row, c, '', ship_fill)
        row += 1

    # ============ CONDICIONES ============
    row += 1
    ws.merge_range(row, 0, row, 12, 'CONDICIONES Y NOTAS', section_fmt)
    row += 1
    notes = [
        'Precios en USD, terminos FOB (o indicar EXW/CIF)',
        'Indicar tiempo de produccion estimado por producto',
        'Indicar peso unitario en gramos si es posible',
        'Productos marcados "Nuevo Odoo" = nuevos en nuestro catalogo',
        'Forma de pago: ___________________________',
        'Observaciones del proveedor: ___________________________',
    ]
    for n in notes:
        ws.write(row, 0, n, note_fmt)
        row += 1

    wb.close()
    for f in tmp_files:
        try: os.unlink(f)
        except: pass

    return filepath

# ==================== GENERAR EXCEL REPOSICION ====================
def generate_reorder_excel(pedido_id):
    conn = get_db()
    pedido = conn.execute("SELECT p.*, pr.nombre as proveedor FROM pedidos p JOIN proveedores pr ON p.proveedor_id=pr.id WHERE p.id=?", (pedido_id,)).fetchone()
    if not pedido:
        conn.close()
        return None, "Pedido no encontrado"

    productos = conn.execute("SELECT * FROM productos WHERE pedido_id=? ORDER BY id", (pedido_id,)).fetchall()
    conn.close()

    filepath = os.path.join(SCRIPT_DIR, f"reposicion_{pedido['proveedor']}_{pedido_id}.xlsx")

    wb = xlsxwriter.Workbook(filepath)
    ws = wb.add_worksheet('Reposicion')

    # Formatos
    header_fmt = wb.add_format({'bold': True, 'bg_color': '#D97706', 'font_color': 'white',
                                 'border': 1, 'text_wrap': True, 'valign': 'vcenter'})
    cell_fmt = wb.add_format({'border': 1, 'valign': 'vcenter', 'text_wrap': True})
    money_fmt = wb.add_format({'border': 1, 'valign': 'vcenter', 'num_format': '$#,##0.00'})
    num_fmt = wb.add_format({'border': 1, 'valign': 'vcenter', 'num_format': '#,##0'})

    # Formatos extra
    total_money_fmt = wb.add_format({'bold': True, 'border': 1, 'valign': 'vcenter',
                                      'num_format': '$#,##0.00', 'bg_color': '#FEF3C7'})
    total_num_fmt = wb.add_format({'bold': True, 'border': 1, 'valign': 'vcenter',
                                    'num_format': '#,##0', 'bg_color': '#FEF3C7'})
    bold_r_fmt = wb.add_format({'bold': True, 'align': 'right', 'border': 1, 'valign': 'vcenter'})
    subtotal_fmt = wb.add_format({'border': 1, 'valign': 'vcenter', 'num_format': '$#,##0.00',
                                   'bg_color': '#FFFBEB'})
    weight_fmt = wb.add_format({'border': 1, 'valign': 'vcenter', 'num_format': '#,##0', 'align': 'center'})

    url_fmt = wb.add_format({'border': 1, 'valign': 'vcenter', 'font_color': '#2563EB',
                              'underline': True, 'font_size': 10})

    # Col: A=Img B=Ref C=Producto D=Variante E=URL F=PrecioFOB G=Cant H=Subtotal I=Peso(g) J=PesoTotal(g) K=LandedUnit L=PVP M=StockOdoo N=Vendido O=Reponer
    headers = ['', 'Ref', 'Producto', 'Variante', 'URL Producto', 'Precio FOB', 'Ult. Cant', 'Subtotal',
               'Peso unit.(g)', 'Peso Total(g)', 'Landed Unit', 'PVP', 'Stock Odoo', 'Vendido 12m', 'Reponer']
    ws.set_row(0, 30)
    for i, h in enumerate(headers):
        ws.write(0, i, h, header_fmt)

    ws.set_column(0, 0, 10)   # imagen
    ws.set_column(1, 1, 10)   # ref
    ws.set_column(2, 2, 35)   # nombre
    ws.set_column(3, 3, 15)   # variante
    ws.set_column(4, 4, 30)   # url
    ws.set_column(5, 5, 12)   # precio
    ws.set_column(6, 6, 10)   # cant
    ws.set_column(7, 7, 12)   # subtotal
    ws.set_column(8, 8, 11)   # peso unit
    ws.set_column(9, 9, 12)   # peso total
    ws.set_column(10, 10, 12)  # landed
    ws.set_column(11, 11, 12)  # pvp
    ws.set_column(12, 12, 10)  # stock
    ws.set_column(13, 13, 12)  # vendido
    ws.set_column(14, 14, 10)  # reponer

    row = 1
    tmp_files = []
    for p in productos:
        ws.set_row(row, 55)

        # Imagen — escalar para que quepa en celda
        if p['imagen_b64'] and p['imagen_b64'].startswith('data:'):
            try:
                b64_data = p['imagen_b64'].split(',', 1)[1]
                img_bytes = base64.b64decode(b64_data)
                tmp = tempfile.NamedTemporaryFile(suffix='.png', delete=False)
                tmp.write(img_bytes)
                tmp.close()
                tmp_files.append(tmp.name)

                pil_img = PILImage.open(io.BytesIO(img_bytes))
                img_w, img_h = pil_img.size
                pil_img.close()
                target_w, target_h = 68, 51  # col 10 units ≈ 70px, row 55px - margins
                scale_x = target_w / img_w if img_w > target_w else 1.0
                scale_y = target_h / img_h if img_h > target_h else 1.0
                scale = min(scale_x, scale_y)

                ws.insert_image(row, 0, tmp.name, {
                    'x_scale': scale, 'y_scale': scale,
                    'x_offset': 2, 'y_offset': 2,
                    'object_position': 1
                })
            except:
                pass

        ws.write(row, 1, p['referencia'], cell_fmt)
        ws.write(row, 2, p['nombre'], cell_fmt)
        ws.write(row, 3, p['variante'], cell_fmt)
        # URL del producto
        prod_url = p['url_proveedor'] or p['url_paracarpinteros'] or ''
        if prod_url:
            ws.write_url(row, 4, prod_url, url_fmt, 'Ver producto')
        else:
            ws.write(row, 4, '', cell_fmt)
        ws.write(row, 5, p['precio'], money_fmt)
        ws.write(row, 6, p['cantidad'], num_fmt)
        # Subtotal = precio * cantidad
        xl_r = row + 1
        ws.write_formula(row, 7, f'=F{xl_r}*G{xl_r}', subtotal_fmt)
        # Peso unitario y total
        peso_g = p['peso_g'] if p['peso_g'] else 0
        ws.write(row, 8, peso_g, weight_fmt)
        ws.write_formula(row, 9, f'=G{xl_r}*I{xl_r}', weight_fmt)
        ws.write(row, 10, p['landed_unit'], money_fmt)
        ws.write(row, 11, p['pvp_unit'], money_fmt)
        ws.write(row, 12, '', cell_fmt)  # stock odoo
        ws.write(row, 13, '', cell_fmt)  # vendido
        ws.write(row, 14, '', num_fmt)   # reponer
        row += 1

    # ===== FILA DE TOTALES =====
    first = 2  # Excel row 2 (1-based, first data row)
    last = row  # 0-based last+1, but in Excel 1-based = row
    ws.set_row(row, 25)
    ws.write(row, 4, 'TOTALES:', bold_r_fmt)
    ws.write_formula(row, 5, f'=SUM(F{first}:F{last})', total_money_fmt)   # sum precios
    ws.write_formula(row, 6, f'=SUM(G{first}:G{last})', total_num_fmt)     # sum cant
    ws.write_formula(row, 7, f'=SUM(H{first}:H{last})', total_money_fmt)   # sum subtotal
    ws.write_formula(row, 8, f'=SUM(I{first}:I{last})', total_num_fmt)     # sum peso unit
    ws.write_formula(row, 9, f'=SUM(J{first}:J{last})', total_num_fmt)     # sum peso total
    ws.write_formula(row, 10, f'=SUM(K{first}:K{last})', total_money_fmt)  # sum landed
    ws.write_formula(row, 11, f'=SUM(L{first}:L{last})', total_money_fmt)  # sum pvp

    wb.close()

    # Limpiar temporales
    for f in tmp_files:
        try:
            os.unlink(f)
        except:
            pass

    return filepath, None

# ==================== CATALOGO ====================
def update_catalogo(productos, proveedor_nombre):
    """Actualiza catálogo maestro con productos de un pedido guardado"""
    conn = get_db()
    updated = 0
    for p in productos:
        ref = p.get('referencia', '').strip()
        if not ref:
            continue
        existing = conn.execute("SELECT id, veces_pedido FROM catalogo WHERE referencia=?", (ref,)).fetchone()
        if existing:
            conn.execute("""UPDATE catalogo SET nombre=COALESCE(NULLIF(?,''),(SELECT nombre FROM catalogo WHERE id=?)),
                variante=COALESCE(NULLIF(?,''),variante), imagen_b64=COALESCE(NULLIF(?,''),imagen_b64),
                ultimo_precio=?, ultimo_proveedor=?, pvp_crc=?, peso_g=?, margen=?,
                url_paracarpinteros=COALESCE(NULLIF(?,''),url_paracarpinteros),
                url_proveedor=COALESCE(NULLIF(?,''),url_proveedor),
                veces_pedido=veces_pedido+1, updated_at=datetime('now') WHERE id=?""",
                (p.get('nombre',''), existing['id'], p.get('variante',''), p.get('imgSrc',''),
                 p.get('precio',0), proveedor_nombre, p.get('pvpUnit',0)*525, p.get('pesoG',0), p.get('margen',40),
                 p.get('url_paracarpinteros',''), p.get('url_proveedor',''), existing['id']))
        else:
            conn.execute("""INSERT INTO catalogo (referencia,nombre,variante,imagen_b64,ultimo_precio,
                ultimo_proveedor,pvp_crc,peso_g,margen,url_paracarpinteros,url_proveedor,veces_pedido)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,1)""",
                (ref, p.get('nombre',''), p.get('variante',''), p.get('imgSrc',''),
                 p.get('precio',0), proveedor_nombre, p.get('pvpUnit',0)*525,
                 p.get('pesoG',0), p.get('margen',40),
                 p.get('url_paracarpinteros',''), p.get('url_proveedor','')))
        updated += 1
    conn.commit()
    conn.close()
    print(f"  [CATALOGO] {updated} productos actualizados")
    return updated

def generate_seo_name(nombre, variante='', referencia=''):
    """Genera nombre SEO para producto en Odoo/web"""
    # Limpiar nombre base
    name = nombre.replace('(sin nombre)', '').strip()
    if not name:
        name = referencia
    # Añadir variante si hay
    if variante:
        name = f"{name} {variante}"
    # Capitalizar palabras importantes
    words = name.split()
    seo = ' '.join(w.capitalize() if len(w) > 2 else w for w in words)
    # Añadir referencia al final
    if referencia and referencia.upper() not in seo.upper():
        seo += f" .Ref.{referencia}"
    # Añadir contexto carpintería si no tiene
    lower = seo.lower()
    if not any(kw in lower for kw in ['carpint', 'madera', 'fresa', 'broca', 'sierra', 'torno', 'gubia', 'router']):
        seo += " - Carpintería"
    return seo

def odoo_create_product(odoo, producto):
    """Crea producto borrador (no publicado) en Odoo"""
    try:
        m = odoo['models']
        db, uid, pwd = odoo['db'], odoo['uid'], odoo['password']

        nombre_seo = generate_seo_name(producto.get('nombre',''),
                                        producto.get('variante',''),
                                        producto.get('referencia',''))

        vals = {
            'name': nombre_seo,
            'default_code': producto.get('referencia', ''),
            'list_price': producto.get('pvp_crc', 0),
            'standard_price': producto.get('ultimo_precio', 0),
            'type': 'product',
            'sale_ok': True,
            'purchase_ok': True,
            'website_published': False,  # NO PUBLICADO
            'description_sale': f"Producto importado - {producto.get('nombre','')}",
        }

        # Imagen: convertir base64 a formato Odoo (sin prefijo data:image)
        img = producto.get('imagen_b64', '')
        if img and ',' in img:
            img = img.split(',', 1)[1]
        if img:
            vals['image_1920'] = img

        product_id = m.execute_kw(db, uid, pwd, 'product.template', 'create', [vals])
        print(f"  [ODOO] Producto creado: id={product_id}, nombre='{nombre_seo}'")
        return product_id, None
    except Exception as e:
        print(f"  [ODOO] Error creando producto: {e}")
        return None, str(e)

# ==================== HTTP HANDLER ====================
class CalcHandler(SimpleHTTPRequestHandler):

    def do_GET(self):
        if self.path == '/' or self.path == '':
            self.path = '/calculadora.html'

        filepath = os.path.join(SCRIPT_DIR, self.path.lstrip('/'))
        if os.path.isfile(filepath):
            self.send_response(200)
            ct = 'text/html; charset=utf-8' if filepath.endswith('.html') else \
                 'application/javascript' if filepath.endswith('.js') else \
                 'application/json' if filepath.endswith('.json') else \
                 'image/png' if filepath.endswith('.png') else \
                 'image/jpeg' if filepath.endswith(('.jpg','.jpeg')) else \
                 'image/svg+xml' if filepath.endswith('.svg') else \
                 'application/octet-stream'
            self.send_header('Content-Type', ct)
            self.end_headers()
            with open(filepath, 'rb') as f:
                self.wfile.write(f.read())
            return

        # API endpoints GET
        path = urlparse(self.path).path
        params = parse_qs(urlparse(self.path).query)

        if path == '/api/proveedores':
            self.json_response(self.db_query("SELECT * FROM proveedores ORDER BY nombre"))

        elif path == '/api/pedidos':
            prov_id = params.get('proveedor_id', [None])[0]
            if prov_id:
                self.json_response(self.db_query(
                    "SELECT p.*, pr.nombre as proveedor FROM pedidos p JOIN proveedores pr ON p.proveedor_id=pr.id WHERE p.proveedor_id=? ORDER BY p.fecha DESC", (prov_id,)))
            else:
                self.json_response(self.db_query(
                    "SELECT p.*, pr.nombre as proveedor FROM pedidos p JOIN proveedores pr ON p.proveedor_id=pr.id ORDER BY p.fecha DESC"))

        elif path == '/api/pedido':
            pid = params.get('id', [None])[0]
            if pid:
                pedido = self.db_query_one("SELECT p.*, pr.nombre as proveedor FROM pedidos p JOIN proveedores pr ON p.proveedor_id=pr.id WHERE p.id=?", (pid,))
                productos = self.db_query("SELECT * FROM productos WHERE pedido_id=? ORDER BY id", (pid,))
                self.json_response({'pedido': pedido, 'productos': productos})
            else:
                self.json_response({'error': 'id requerido'}, 400)

        elif path == '/api/odoo/config':
            self.json_response(self.db_query_one("SELECT url, db_name, username, api_key, last_sync FROM odoo_config WHERE id=1"))

        elif path == '/api/odoo/test':
            odoo, err = odoo_connect()
            self.json_response({'ok': odoo is not None, 'error': err})

        elif path == '/api/odoo/stock':
            # Consulta stock+ventas+POs para lista de referencias (via query param refs=P07,G08,...)
            refs_str = params.get('refs', [''])[0]
            refs = [r.strip() for r in refs_str.split(',') if r.strip()]
            if not refs:
                self.json_response({'error': 'refs requerido (ej: ?refs=P07,G08)'}, 400)
                return
            odoo, err = odoo_connect()
            if not odoo:
                self.json_response({'error': err}, 500)
                return
            data, err = odoo_get_stock_and_orders(odoo, refs)
            self.json_response({'data': data, 'error': err})

        elif path == '/api/catalogo':
            q = params.get('q', [''])[0].strip()
            if q:
                like = f'%{q}%'
                rows = self.db_query(
                    "SELECT * FROM catalogo WHERE nombre LIKE ? OR referencia LIKE ? OR variante LIKE ? ORDER BY veces_pedido DESC, updated_at DESC LIMIT 100",
                    (like, like, like))
            else:
                rows = self.db_query("SELECT * FROM catalogo ORDER BY veces_pedido DESC, updated_at DESC LIMIT 200")
            self.json_response(rows)

        elif path == '/api/waitlist':
            rows = self.db_query("""
                SELECT pr.id as producto_id, pr.referencia, pr.nombre, pr.nota, pr.cantidad,
                       pr.cliente_avisado, pr.cliente_completado,
                       p.titulo as pedido_titulo, p.estado, p.id as pedido_id
                FROM productos pr
                JOIN pedidos p ON pr.pedido_id = p.id
                WHERE pr.nota IS NOT NULL AND pr.nota != ''
                ORDER BY pr.cliente_completado ASC, p.estado DESC, pr.nota
            """)
            self.json_response(rows)

        elif path == '/api/reorder':
            pid = params.get('pedido_id', [None])[0]
            if not pid:
                self.json_response({'error': 'pedido_id requerido'}, 400)
                return
            filepath, err = generate_reorder_excel(int(pid))
            if err:
                self.json_response({'error': err}, 500)
                return
            self.send_response(200)
            self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            self.send_header('Content-Disposition', f'attachment; filename="{os.path.basename(filepath)}"')
            self.end_headers()
            with open(filepath, 'rb') as f:
                self.wfile.write(f.read())
            os.unlink(filepath)

        else:
            self.send_error(404)

    def do_POST(self):
        body = self.rfile.read(int(self.headers.get('Content-Length', 0)))
        path = urlparse(self.path).path

        if path == '/api/parse':
            tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
            tmp.write(body)
            tmp.close()
            try:
                headers, rows, images = parse_excel(tmp.name)
                self.json_response({'ok': True, 'headers': headers, 'rows': rows,
                                     'images': {str(k): v for k, v in images.items()},
                                     'imageCount': len(images)})
            except Exception as e:
                self.json_response({'ok': False, 'error': str(e)}, 500)
            finally:
                os.unlink(tmp.name)

        elif path == '/api/upload-excel':
            # Mismo que /api/parse pero acepta multipart/form-data
            try:
                content_type = self.headers.get('Content-Type', '')
                if 'multipart' in content_type:
                    # Extraer boundary y parsear multipart
                    boundary = content_type.split('boundary=')[1].strip()
                    if boundary.startswith('"'):
                        boundary = boundary[1:-1]
                    parts = body.split(('--' + boundary).encode())
                    file_data = None
                    for part in parts:
                        if b'filename=' in part:
                            # Encontrar inicio de datos (después de doble newline)
                            idx = part.find(b'\r\n\r\n')
                            if idx >= 0:
                                file_data = part[idx+4:]
                                # Quitar trailing \r\n--
                                if file_data.endswith(b'\r\n'):
                                    file_data = file_data[:-2]
                                break
                    if not file_data:
                        self.json_response({'ok': False, 'error': 'No file found'}, 400)
                        return
                else:
                    file_data = body

                tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
                tmp.write(file_data)
                tmp.close()
                try:
                    headers, rows, images = parse_excel(tmp.name)
                    self.json_response({'ok': True, 'headers': headers, 'rows': rows,
                                         'images': {str(k): v for k, v in images.items()},
                                         'imageCount': len(images)})
                except Exception as e:
                    self.json_response({'ok': False, 'error': str(e)}, 500)
                finally:
                    os.unlink(tmp.name)
            except Exception as e:
                self.json_response({'ok': False, 'error': str(e)}, 500)

        elif path == '/api/proveedor/save':
            data = json.loads(body)
            conn = get_db()
            if data.get('id'):
                conn.execute("UPDATE proveedores SET nombre=?, contacto=?, notas=? WHERE id=?",
                             (data['nombre'], data.get('contacto', ''), data.get('notas', ''), data['id']))
            else:
                conn.execute("INSERT OR IGNORE INTO proveedores (nombre, contacto, notas) VALUES (?,?,?)",
                             (data['nombre'], data.get('contacto', ''), data.get('notas', '')))
            conn.commit()
            prov = conn.execute("SELECT * FROM proveedores WHERE nombre=?", (data['nombre'],)).fetchone()
            conn.close()
            self.json_response(dict(prov))

        elif path == '/api/pedido/save':
            data = json.loads(body)
            conn = get_db()

            # Asegurar proveedor existe
            prov = conn.execute("SELECT id FROM proveedores WHERE nombre=?", (data['proveedor'],)).fetchone()
            if not prov:
                conn.execute("INSERT INTO proveedores (nombre) VALUES (?)", (data['proveedor'],))
                conn.commit()
                prov = conn.execute("SELECT id FROM proveedores WHERE nombre=?", (data['proveedor'],)).fetchone()

            prov_id = prov['id']
            ped = data.get('pedido', {})

            if ped.get('id'):
                conn.execute("""UPDATE pedidos SET titulo=?, estado=?, notas=?, flete=?, seguro=?,
                    gastos_json=?, impuestos_json=?, alloc_peso=?, fx_usd_crc=? WHERE id=?""",
                    (ped.get('titulo', ''), ped.get('estado', 'borrador'), ped.get('notas', ''),
                     ped.get('flete', 0), ped.get('seguro', 0),
                     json.dumps(ped.get('gastos', [])), json.dumps(ped.get('impuestos', [])),
                     ped.get('alloc_peso', 50), ped.get('fx_usd_crc', 525), ped['id']))
                pedido_id = ped['id']
                conn.execute("DELETE FROM productos WHERE pedido_id=?", (pedido_id,))
            else:
                cur = conn.execute("""INSERT INTO pedidos (proveedor_id, titulo, estado, notas, flete, seguro,
                    gastos_json, impuestos_json, alloc_peso, fx_usd_crc) VALUES (?,?,?,?,?,?,?,?,?,?)""",
                    (prov_id, ped.get('titulo', data['proveedor'] + ' ' + datetime.now().strftime('%Y-%m-%d')),
                     ped.get('estado', 'borrador'), ped.get('notas', ''),
                     ped.get('flete', 0), ped.get('seguro', 0),
                     json.dumps(ped.get('gastos', [])), json.dumps(ped.get('impuestos', [])),
                     ped.get('alloc_peso', 50), ped.get('fx_usd_crc', 525)))
                pedido_id = cur.lastrowid

            # Guardar productos
            for p in data.get('productos', []):
                conn.execute("""INSERT INTO productos (pedido_id, referencia, nombre, variante,
                    url_proveedor, url_paracarpinteros, imagen_b64, moneda, precio, cantidad,
                    peso_g, vol_cm3, margen, landed_unit, pvp_unit, nota) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (pedido_id, p.get('referencia', ''), p.get('nombre', ''), p.get('variante', ''),
                     p.get('url_proveedor', ''), p.get('url_paracarpinteros', ''),
                     p.get('imgSrc', ''), p.get('currency', 'USD'),
                     p.get('precio', 0), p.get('cant', 0), p.get('pesoG', 0),
                     p.get('volCm3', 0), p.get('margen', 40),
                     p.get('landedUnit', 0), p.get('pvpUnit', 0), p.get('nota', '')))

            conn.commit()
            conn.close()

            # Auto-alimentar catálogo
            try:
                update_catalogo(data.get('productos', []), data['proveedor'])
            except Exception as e:
                print(f"  [CATALOGO] Error: {e}")

            self.json_response({'ok': True, 'pedido_id': pedido_id})

        elif path == '/api/pedido/update-estado':
            data = json.loads(body)
            conn = get_db()
            conn.execute("UPDATE pedidos SET estado=? WHERE id=?", (data['estado'], data['id']))
            conn.commit()
            conn.close()
            self.json_response({'ok': True})

        elif path == '/api/waitlist/update':
            data = json.loads(body)
            pid = data.get('producto_id')
            conn = get_db()
            if 'avisado' in data:
                conn.execute("UPDATE productos SET cliente_avisado=? WHERE id=?", (1 if data['avisado'] else 0, pid))
            if 'completado' in data:
                conn.execute("UPDATE productos SET cliente_completado=? WHERE id=?", (1 if data['completado'] else 0, pid))
            conn.commit()
            conn.close()
            self.json_response({'ok': True})

        elif path == '/api/pedido/delete':
            data = json.loads(body)
            conn = get_db()
            conn.execute("DELETE FROM productos WHERE pedido_id=?", (data['id'],))
            conn.execute("DELETE FROM pedidos WHERE id=?", (data['id'],))
            conn.commit()
            conn.close()
            self.json_response({'ok': True})

        elif path == '/api/odoo/config':
            try:
                data = json.loads(body)
                print(f"  [SAVE ODOO] url='{data.get('url','')}' db='{data.get('db_name','')}' user='{data.get('username','')}' key={'***'+data.get('api_key','')[-4:] if data.get('api_key') else 'VACIA'}")
                conn = get_db()
                conn.execute("UPDATE odoo_config SET url=?, db_name=?, username=?, api_key=? WHERE id=1",
                             (data.get('url', ''), data.get('db_name', ''), data.get('username', ''), data.get('api_key', '')))
                conn.commit()
                check = conn.execute("SELECT url, api_key FROM odoo_config WHERE id=1").fetchone()
                conn.close()
                saved_ok = bool(check and check['url'] and check['api_key'])
                print(f"  [SAVE ODOO] Verificacion: {'OK' if saved_ok else 'FALLO'}")
                self.json_response({'ok': saved_ok, 'verified': saved_ok})
            except Exception as e:
                print(f"  [SAVE ODOO] ERROR: {e}")
                self.json_response({'ok': False, 'error': str(e)}, 500)

        elif path == '/api/scrape':
            try:
                data = json.loads(body)
                urls = data.get('urls', [])
                if not urls:
                    self.json_response({'ok': True, 'results': {}})
                    return
                results = scrape_multiple_urls(urls)
                # Convertir a formato serializable
                out = {}
                for url, info in results.items():
                    out[url] = {
                        'nombre': info.get('nombre', ''),
                        'referencia': info.get('referencia', ''),
                        'sku': info.get('sku', ''),
                        'imagen_b64': info.get('imagen_b64', ''),
                        'precio_crc': info.get('precio_crc', 0),
                        'categoria': info.get('categoria', ''),
                    }
                self.json_response({'ok': True, 'results': out})
            except Exception as e:
                print(f"  [SCRAPE] Error endpoint: {e}")
                self.json_response({'ok': False, 'error': str(e)}, 500)

        elif path == '/api/scrape-titles':
            try:
                data = json.loads(body)
                urls = data.get('urls', [])  # [{url, index}]
                results = {}
                for item in urls:
                    url = item.get('url', '')
                    idx = item.get('index', 0)
                    title = scrape_supplier_title(url)
                    if title:
                        results[str(idx)] = title
                self.json_response({'ok': True, 'titles': results})
            except Exception as e:
                print(f"  [TITLES] Error: {e}")
                self.json_response({'ok': False, 'error': str(e)}, 500)

        elif path == '/api/supplier-template':
            try:
                data = json.loads(body)
                filepath = generate_supplier_template(data.get('productos', []))
                self.send_response(200)
                self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                self.send_header('Content-Disposition', 'attachment; filename="cotizacion_proveedor.xlsx"')
                self.end_headers()
                with open(filepath, 'rb') as f:
                    self.wfile.write(f.read())
                os.unlink(filepath)
            except Exception as e:
                self.json_response({'error': str(e)}, 500)

        elif path == '/api/catalogo/to-order':
            # Recibe lista de IDs del catálogo, los carga como productos en la calculadora
            data = json.loads(body)
            ids = data.get('ids', [])
            if not ids:
                self.json_response({'ok': False, 'error': 'No hay productos seleccionados'}, 400)
                return
            conn = get_db()
            placeholders = ','.join('?' * len(ids))
            rows = conn.execute(f"SELECT * FROM catalogo WHERE id IN ({placeholders})", ids).fetchall()
            conn.close()
            productos = []
            for r in rows:
                r = dict(r)
                productos.append({
                    'referencia': r['referencia'], 'nombre': r['nombre'], 'variante': r['variante'],
                    'currency': 'USD', 'precio': r['ultimo_precio'], 'cant': 1,
                    'pesoG': r['peso_g'], 'volCm3': 0, 'margen': r['margen'],
                    'imgSrc': r['imagen_b64'], 'url_proveedor': r['url_proveedor'],
                    'url_paracarpinteros': r['url_paracarpinteros']
                })
            self.json_response({'ok': True, 'productos': productos})

        elif path == '/api/odoo/create-product':
            try:
                data = json.loads(body)
                odoo, err = odoo_connect()
                if not odoo:
                    self.json_response({'ok': False, 'error': err}, 500)
                    return
                product_id, err = odoo_create_product(odoo, data)
                if err:
                    self.json_response({'ok': False, 'error': err}, 500)
                    return
                # Actualizar catálogo con el ID de Odoo
                if data.get('catalogo_id'):
                    conn = get_db()
                    conn.execute("UPDATE catalogo SET odoo_product_id=?, odoo_publicado=0 WHERE id=?",
                                 (product_id, data['catalogo_id']))
                    conn.commit()
                    conn.close()
                self.json_response({'ok': True, 'product_id': product_id})
            except Exception as e:
                self.json_response({'ok': False, 'error': str(e)}, 500)

        else:
            self.send_error(404)

    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'POST, GET, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.end_headers()

    def json_response(self, data, code=200):
        self.send_response(code)
        self.send_header('Content-Type', 'application/json; charset=utf-8')
        self.send_header('Access-Control-Allow-Origin', '*')
        self.end_headers()
        if isinstance(data, sqlite3.Row):
            data = dict(data)
        elif isinstance(data, list) and data and isinstance(data[0], sqlite3.Row):
            data = [dict(r) for r in data]
        self.wfile.write(json.dumps(data, ensure_ascii=False, default=str).encode('utf-8'))

    def db_query(self, sql, params=()):
        conn = get_db()
        rows = conn.execute(sql, params).fetchall()
        conn.close()
        return [dict(r) for r in rows]

    def db_query_one(self, sql, params=()):
        conn = get_db()
        row = conn.execute(sql, params).fetchone()
        conn.close()
        return dict(row) if row else None

    def log_message(self, format, *args):
        msg = format % args
        if '/api/' in msg:
            print(f"  {msg}")

# ==================== MAIN ====================
def main():
    print("=" * 52)
    print("  CALCULADORA IMPORTACION v2 - Paracarpinteros")
    print("=" * 52)
    print()

    init_db()

    server = HTTPServer(('0.0.0.0', PORT), CalcHandler)

    # Mostrar IP local para acceso desde otros equipos
    local_ip = '?'
    try:
        import socket
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(('8.8.8.8', 80))
        local_ip = s.getsockname()[0]
        s.close()
    except:
        pass

    print(f"  Local:    http://localhost:{PORT}")
    print(f"  Red:      http://{local_ip}:{PORT}  (otros equipos)")
    print(f"  Carpeta:  {SCRIPT_DIR}")
    print(f"  DB:       {DB_PATH}")
    print()
    print("  NO CIERRES esta ventana mientras uses la calculadora")
    print()

    # webbrowser disabled (Docker mode)

    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\n  Servidor detenido.")
        server.server_close()

if __name__ == '__main__':
    main()
